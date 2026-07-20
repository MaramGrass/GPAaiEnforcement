"""
convert.py — DPA Tracker Excel → JSON converter
------------------------------------------------
Run this whenever you update dpa_tracker_template.xlsx:

    python convert.py

It will produce dpa-data.json in the same folder.
Upload dpa-data.json to your server alongside the app.
"""

import json
from openpyxl import load_workbook

EXCEL_FILE = "dpa_tracker_template.xlsx"
JSON_FILE  = "dpa-data.json"
SHEET_NAME = "DPA Enforcement Actions"

wb = load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]

headers = [cell.value for cell in ws[1]]
rows = []

for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    record = dict(zip(headers, row))
    if not record.get("dpa"):
        continue
    rows.append({
        "id":          str(record.get("id") or f"row-{len(rows)}"),
        "dpa":         str(record.get("dpa") or ""),
        "country":     str(record.get("country") or ""),
        "countryCode": str(record.get("countryCode") or ""),
        "date":        str(record.get("date") or ""),
        "subject":     str(record.get("subject") or ""),
        "summary":     str(record.get("summary") or ""),
        "fine":        str(record["fine"]) if record.get("fine") else None,
        "fineUSD":     float(record["fineUSD"]) if record.get("fineUSD") else None,
        "outcome":     str(record.get("outcome") or "Other"),
        "reference":   str(record.get("reference") or ""),
        "tags":        [t.strip() for t in str(record.get("tags") or "").split(",") if t.strip()],
    })

with open(JSON_FILE, "w", encoding="utf-8") as f:
    json.dump(rows, f, indent=2, ensure_ascii=False)

print(f"✓ Exported {len(rows)} cases to {JSON_FILE}")
